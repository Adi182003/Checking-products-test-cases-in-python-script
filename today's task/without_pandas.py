#libraries required
from openpyxl import load_workbook
from collections import defaultdict
#without_pandas.py              
def get_top_products():
    """
    Loads data from an Excel file, processes it to find the top 5 products
    by quantity, and returns the result.
    """
    #loading excel file and parsing sheets
    wb = load_workbook("Data For Test.xlsx", data_only=True)
    transaction_sheet = wb["Transaction"]
    item_info_sheet = wb["Item_Info"]

    # Extracting data from the sheets
    transaction_data = list(transaction_sheet.iter_rows(values_only=True))
    item_info_data = list(item_info_sheet.iter_rows(values_only=True))

    # Creating a mapping of Material No to Product Head
    transaction_headers = transaction_data[0]
    item_info_headers = item_info_data[0]

    transaction_idx = {header: idx for idx, header in enumerate(transaction_headers)}
    item_info_idx = {header: idx for idx, header in enumerate(item_info_headers)}

    material_to_product = {
        row[item_info_idx["Material No"]]: row[item_info_idx["Product Head"]]
        for row in item_info_data[1:]
    }

    # Calculating total quantities for each product
    product_quantities = defaultdict(int)

    for row in transaction_data[1:]:
        material_no = row[transaction_idx["Material No"]]
        qty = row[transaction_idx["Qty"]]
        product_head = material_to_product.get(material_no)
        if isinstance(qty, (int, float)):
            product_quantities[(product_head, material_no)] += qty

    # Sorting and selecting the top 5 products
    top_5 = sorted(product_quantities.items(), key=lambda x: x[1], reverse=True)[:5]

    # Formatting output
    output = [
        {"Product Head": k[0], "Material No": k[1], "Qty": v}
        for k, v in top_5
    ]
    return output

if __name__ == "__main__":
    from tabulate import tabulate
    top_products = get_top_products()
    #printing the results
    print("\nTop 5 Products by Quantity:\n")
    print(tabulate(top_products, headers="keys", tablefmt="fancy_grid"))
